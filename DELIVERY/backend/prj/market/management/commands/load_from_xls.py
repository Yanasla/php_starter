from django.core.management.base import BaseCommand, CommandError

from market.models import Category, SubCategory, Product

from prj.settings import DATA_DIR

from openpyxl import load_workbook

class Command(BaseCommand):
    
    def handle(self, *args, **options):
        print('Clear DB')
        Category.objects.all().delete()
        SubCategory.objects.all().delete()
        Product.objects.all().delete()

        print('Start importing from excel %s'% DATA_DIR)

        wb = load_workbook('/home/user/Рабочий стол/php_starter/DELIVERY/backend/prj/market/init_data'+'/price.xlsx')
        #print(wb.get_sheet_names())
        sheet = wb.get_sheet_by_name(wb.get_sheet_names()[0])

        print(sheet.max_row)
        cat = None
        for cnt in range(1,sheet.max_row+1):
            item = sheet.cell(row=cnt, column=4).value
            sub_id = sheet.cell(row=cnt, column=2).value
            id = sheet.cell(row=cnt, column=1).value
            if id != None:
                print('Create new Category')
                cat = Category()
                cat.name = id
                cat.save()
            if sub_id != None:
                print('Create new SubCategory')
                scat = SubCategory()
                scat.name = sub_id
                scat.category = cat
                scat.save()
            if item != None:
                print('Create new Good')
                good = Product()
                good.name = item
                good.category = scat
                good.save()



